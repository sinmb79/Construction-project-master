from pathlib import Path

from openpyxl import load_workbook

from civilplan_mcp.tools.boq_generator import generate_boq_excel
from civilplan_mcp.tools.doc_generator import generate_investment_doc
from civilplan_mcp.tools.drawing_generator import generate_svg_drawing
from civilplan_mcp.tools.quantity_estimator import estimate_quantities
from civilplan_mcp.tools.project_parser import parse_project
from civilplan_mcp.tools.unit_price_query import get_unit_prices


def _sample_inputs(tmp_path: Path) -> tuple[dict, dict, dict]:
    project_spec = parse_project(
        description="소로 신설 L=890m B=6m 아스콘 2차선 상하수도 경기도 둔턱지역 2026~2028"
    )
    project_spec["output_dir"] = str(tmp_path)
    quantities = estimate_quantities(
        road_class="소로",
        length_m=890,
        width_m=6.0,
        terrain="구릉",
        pavement_type="아스콘",
        include_water_supply=True,
        include_sewage=True,
        include_retaining_wall=True,
        include_bridge=False,
        bridge_length_m=0.0,
    )
    unit_prices = get_unit_prices(region="경기도", year=2026)
    return project_spec, quantities, unit_prices


def test_generate_boq_excel_creates_workbook(tmp_path: Path) -> None:
    project_spec, quantities, unit_prices = _sample_inputs(tmp_path)

    result = generate_boq_excel(
        project_name="소로 개설(신설) 공사",
        project_spec=project_spec,
        quantities=quantities,
        unit_prices=unit_prices,
        region="경기도",
        year=2026,
        output_filename="boq.xlsx",
    )

    workbook = load_workbook(result["file_path"], data_only=False)
    assert "사업내역서(BOQ)" in workbook.sheetnames


def test_generate_investment_doc_creates_docx(tmp_path: Path) -> None:
    project_spec, quantities, unit_prices = _sample_inputs(tmp_path)
    boq = generate_boq_excel(
        project_name="소로 개설(신설) 공사",
        project_spec=project_spec,
        quantities=quantities,
        unit_prices=unit_prices,
        region="경기도",
        year=2026,
        output_filename="boq_for_doc.xlsx",
    )

    result = generate_investment_doc(
        project_name="소로 개설(신설) 공사",
        project_spec=project_spec,
        quantities=quantities,
        legal_procedures={"summary": {"total_procedures": 3}, "phases": {}},
        boq_summary=boq["summary"],
        requester="22B Labs",
        output_filename="investment.docx",
    )

    assert Path(result["file_path"]).exists()


def test_generate_svg_drawing_creates_svg(tmp_path: Path) -> None:
    project_spec, quantities, _ = _sample_inputs(tmp_path)

    result = generate_svg_drawing(
        drawing_type="횡단면도",
        project_spec=project_spec,
        quantities=quantities,
        scale="1:200",
        output_filename="road.svg",
    )

    content = Path(result["file_path"]).read_text(encoding="utf-8")
    assert "<svg" in content
