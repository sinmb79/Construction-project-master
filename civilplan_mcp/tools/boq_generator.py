from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from civilplan_mcp.config import get_settings
from civilplan_mcp.models import ProjectDomain
from civilplan_mcp.tools._base import wrap_response


def _resolve_output_dir(project_spec: dict[str, Any]) -> Path:
    if project_spec.get("output_dir"):
        path = Path(project_spec["output_dir"])
        path.mkdir(parents=True, exist_ok=True)
        return path
    return get_settings().output_dir


def _price_lookup(unit_prices: dict[str, Any]) -> dict[str, int]:
    return {item["item"]: item["adjusted_price"] for item in unit_prices["results"]}


def generate_boq_excel(
    *,
    project_name: str,
    project_spec: dict[str, Any],
    quantities: dict[str, Any],
    unit_prices: dict[str, Any],
    region: str,
    year: int,
    output_filename: str | None = None,
) -> dict[str, Any]:
    output_dir = _resolve_output_dir(project_spec)
    path = output_dir / (output_filename or f"{project_name}_{year}.xlsx")
    workbook = Workbook()
    overview = workbook.active
    overview.title = "사업개요"
    boq_sheet = workbook.create_sheet("사업내역서(BOQ)")
    workbook.create_sheet("물량산출근거")
    workbook.create_sheet("간접비산출")
    workbook.create_sheet("총사업비요약")
    workbook.create_sheet("연도별투자계획")

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    overview["A1"] = "CivilPlan BOQ"
    overview["A2"] = project_name
    overview["A3"] = "본 자료는 개략 검토용(±20~30% 오차)으로 공식 발주·계약·제출에 사용 불가합니다."

    boq_sheet.append(["공종", "항목", "수량", "단위단가", "금액"])
    for cell in boq_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    price_map = _price_lookup(unit_prices)
    direct_cost = 0
    for category, items in quantities["quantities"].items():
        for item_name, quantity in items.items():
            mapped_price = next((price for name, price in price_map.items() if name.split("(")[0] in item_name or item_name in name), 50000)
            amount = round(float(quantity) * mapped_price)
            direct_cost += amount
            boq_sheet.append([category, item_name, quantity, mapped_price, amount])

    indirect_cost = round(direct_cost * 0.185)
    total_cost = direct_cost + indirect_cost

    summary_sheet = workbook["총사업비요약"]
    summary_sheet.append(["직접공사비", direct_cost])
    summary_sheet.append(["간접비", indirect_cost])
    summary_sheet.append(["총사업비", total_cost])

    plan_sheet = workbook["연도별투자계획"]
    start = project_spec.get("year_start", year)
    end = project_spec.get("year_end", year)
    years = list(range(start, end + 1))
    default_ratios = [0.3, 0.5, 0.2][: len(years)] or [1.0]
    if len(default_ratios) < len(years):
        default_ratios = [round(1 / len(years), 2) for _ in years]
    for target_year, ratio in zip(years, default_ratios):
        plan_sheet.append([target_year, round(total_cost * ratio)])

    workbook.save(path)

    return wrap_response(
        {
            "status": "success",
            "file_path": str(path),
            "summary": {
                "direct_cost": direct_cost,
                "indirect_cost": indirect_cost,
                "total_cost": total_cost,
                "total_cost_billion": round(total_cost / 100000000, 2),
                "sheets": workbook.sheetnames,
                "region": region,
                "year": year,
            },
        },
        ProjectDomain.복합,
    )
